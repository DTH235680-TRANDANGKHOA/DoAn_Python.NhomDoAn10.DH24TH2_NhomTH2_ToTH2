import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from db import connect_db
from datetime import date
import openpyxl 

current_cart = [] # (ma_tivi, ten_tivi, so_luong, don_gia, thanh_tien) # Giỏ hàng hiện tại
tivi_list = [] # (ma_tivi, ten_tivi, don_gia, so_luong_ton) # Danh sách Tivi còn hàng
customer_list = [] # (ma_kh, ho_ten) # Danh sách Khách hàng
staff_list = [] # (ma_nv, ho_ten) # Danh sách Nhân viên

# Biến toàn cục để lưu thông tin hóa đơn vừa được tạo
last_invoice_info = None # (ma_hd, ma_kh, ma_nv, tong_tien, current_cart_data)

def create_view(parent_frame):
    global current_cart, tivi_list, customer_list, staff_list, last_invoice_info
    
    # Reset giỏ hàng và thông tin hóa đơn cuối cùng mỗi khi vào tab
    current_cart = []
    last_invoice_info = None 
    
    # Tạo khung chính
    frame = tk.Frame(parent_frame, bg="#f4f4f4")
    frame.pack(fill="both", expand=True)
    #--- Tiêu đề ---
    tk.Label(frame, text="BÁN HÀNG - HÓA ĐƠN", font=("Arial", 20, "bold"), bg="#f4f4f4").pack(pady=20) 

    # --- Khung chính chứa các thành phần theo bố cục Grid ---
    main_container = tk.Frame(frame, bg="#f4f4f4")
    main_container.pack(fill="both", expand=True, padx=20, pady=10)
    
    # Cấu hình Grid cho main_container
    main_container.columnconfigure(0, weight=1) 
    main_container.columnconfigure(1, weight=1) 
    main_container.rowconfigure(1, weight=1) 

    # --- HÀNG 1 (Grid Row 0) ---
    
    # 1. CỘT TRÁI (Thông tin Hóa đơn)
    customer_frame = tk.LabelFrame(main_container, text="Thông tin Hóa đơn", padx=10, pady=10)
    customer_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=5)

    # -- Combobox Khách hàng và Nhân viên ---
    tk.Label(customer_frame, text="Chọn Khách hàng:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
    customer_combo = ttk.Combobox(customer_frame, width=30, state="readonly")
    customer_combo.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    
    tk.Label(customer_frame, text="Chọn Nhân viên:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
    staff_combo = ttk.Combobox(customer_frame, width=30, state="readonly")
    staff_combo.grid(row=1, column=1, sticky="w", padx=5, pady=5)
    
    customer_frame.grid_columnconfigure(1, weight=1)


    # 2. CỘT PHẢI (Thêm Tivi vào giỏ)
    product_frame = tk.LabelFrame(main_container, text="Thêm Tivi vào giỏ", padx=10, pady=10)
    product_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=5)
    
    #-- Combobox Tivi và nhập số lượng ---
    tk.Label(product_frame, text="Chọn Tivi:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
    tivi_combo = ttk.Combobox(product_frame, width=45, state="readonly")
    tivi_combo.grid(row=0, column=1, columnspan=3, sticky="ew", padx=5, pady=5)
    
    #-- Nhãn hiển thị số lượng tồn và đơn giá ---
    tk.Label(product_frame, text="Số lượng tồn:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
    ton_kho_label = tk.Label(product_frame, text="...", fg="blue")
    ton_kho_label.grid(row=1, column=1, sticky="w", padx=5, pady=5)
    
    tk.Label(product_frame, text="Đơn giá:").grid(row=1, column=2, sticky="w", padx=5, pady=5)
    don_gia_label = tk.Label(product_frame, text="...", fg="blue")
    don_gia_label.grid(row=1, column=3, sticky="w", padx=5, pady=5)
    
    #-- Nhập số lượng mua ---
    tk.Label(product_frame, text="Nhập Số lượng mua:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
    so_luong_entry = tk.Entry(product_frame, width=10)
    so_luong_entry.grid(row=2, column=1, sticky="w", padx=5, pady=5)
    
    # -- Nút Thêm vào giỏ ---
    add_to_cart_btn = tk.Button(product_frame, text="Thêm vào giỏ", command=lambda: add_to_cart(tivi_combo, so_luong_entry, ton_kho_label, don_gia_label, cart_tree, total_label))
    add_to_cart_btn.grid(row=2, column=2, columnspan=2, sticky="ew", padx=5, pady=10)
    
    product_frame.grid_columnconfigure(1, weight=1)
    product_frame.grid_columnconfigure(3, weight=1)


    # --- HÀNG 2 (Grid Row 1) - Giỏ hàng và Thanh toán ---
    cart_frame_container = tk.Frame(main_container, bg="#f4f4f4")
    cart_frame_container.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(10, 0))
    
    # Giỏ hàng
    cart_frame = tk.LabelFrame(cart_frame_container, text="Giỏ hàng hiện tại", padx=10, pady=10)
    cart_frame.pack(fill="both", expand=True)
    #-- TreeView Giỏ hàng ---
    cart_cols = ('ma_tivi', 'ten_tivi', 'so_luong', 'don_gia', 'thanh_tien')
    cart_tree = ttk.Treeview(cart_frame, columns=cart_cols, show='headings', height=10)
    # Định nghĩa cột
    cart_tree.heading('ma_tivi', text='Mã TV')
    cart_tree.heading('ten_tivi', text='Tên Tivi')
    cart_tree.heading('so_luong', text='SL')
    cart_tree.heading('don_gia', text='Đơn giá')
    cart_tree.heading('thanh_tien', text='Thành tiền')
    # Định nghĩa kích thước cột
    cart_tree.column('ma_tivi', width=70, anchor="center")
    cart_tree.column('ten_tivi', width=300, anchor="w")
    cart_tree.column('so_luong', width=70, anchor="center")
    cart_tree.column('don_gia', width=150, anchor="e")
    cart_tree.column('thanh_tien', width=150, anchor="e")
    # Đặt TreeView vào frame
    cart_tree.pack(fill="both", expand=True, side="top")
    
    #  Nút Xóa khỏi giỏ
    remove_item_btn = tk.Button(cart_frame_container, text="Xóa Tivi khỏi giỏ", command=lambda: remove_from_cart(cart_tree, total_label))
    remove_item_btn.pack(pady=5, side="top", anchor="e") # Căn nút Xóa qua phải

    # Thanh toán
    total_frame = tk.Frame(cart_frame_container, bg="#f4f4f4")
    total_frame.pack(fill="x", side="bottom", pady=10)

    #-- Nhãn Tổng tiền và Nút Thanh toán ---
    total_frame.grid_columnconfigure(0, weight=1) 
    total_frame.grid_columnconfigure(1, weight=1) 
    
    # Nút Xuất Excel ĐỘC LẬP
    export_last_btn = tk.Button(total_frame, text="Xuất Hóa Đơn Vừa Lập ra Excel", 
                                font=("Arial", 10), bg="#ffc107", fg="black", 
                                command=lambda: export_last_invoice_from_variable(export_last_btn))
    export_last_btn.grid(row=0, column=0, sticky="w", padx=10, pady=(0, 5)) # Đặt ở góc trái trên
    # Khởi tạo nút xuất Excel bị Disabled
    export_last_btn.config(state=tk.DISABLED)

    tk.Label(total_frame, text="TỔNG TIỀN:", font=("Arial", 14, "bold"), bg="#f4f4f4").grid(row=1, column=0, sticky="w", padx=10)
    total_label = tk.Label(total_frame, text="0 VNĐ", font=("Arial", 14, "bold"), fg="red", bg="#f4f4f4")
    total_label.grid(row=1, column=1, sticky="e", padx=10)
    
    checkout_btn = tk.Button(total_frame, text="THANH TOÁN (LƯU HÓA ĐƠN)", font=("Arial", 12, "bold"), bg="#28a745", fg="white", 
                             command=lambda: checkout(customer_combo, staff_combo, cart_tree, total_label, export_last_btn))
    checkout_btn.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=(10, 0), ipady=10)


    # --- CÁC HÀM XỬ LÝ CHỨC NĂNG ---

    def export_invoice_to_excel(ma_hd_moi, ma_kh, ma_nv, tong_tien, current_cart_data):
        # Lấy tên khách hàng và nhân viên từ list toàn cục
        khach_hang_ten = next((kh[1] for kh in customer_list if kh[0] == ma_kh), f"KH_ID_{ma_kh}")
        nhan_vien_ten = next((nv[1] for nv in staff_list if nv[0] == ma_nv), f"NV_ID_{ma_nv}")
        
        # Dữ liệu Header
        hoadon_info = (ma_hd_moi, date.today(), nhan_vien_ten, khach_hang_ten, tong_tien)
        
        # Mở hộp thoại lưu file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"HoaDon_{ma_hd_moi}_{khach_hang_ten.replace(' ', '')}.xlsx"
        )

        if not file_path:
            return # Người dùng hủy

        try:
            # 1. Tạo Workbook và Sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"HD_{ma_hd_moi}"
            
            # 2. Ghi thông tin Hóa đơn (Phần Header)
            ws.merge_cells('A1:E1')
            ws['A1'] = "THÔNG TIN HÓA ĐƠN"
            ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
            ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            ws['A3'] = "Mã HĐ:"
            ws['B3'] = hoadon_info[0] 
            ws['A4'] = "Ngày Lập:"
            ws['B4'] = hoadon_info[1] 
            ws['A5'] = "Nhân Viên:"
            ws['B5'] = hoadon_info[2] 
            ws['C3'] = "Khách Hàng:"
            ws['D3'] = hoadon_info[3] 
            
            ws['C5'] = "TỔNG TIỀN THANH TOÁN:"
            ws['D5'] = hoadon_info[4] 
            ws['D5'].font = openpyxl.styles.Font(bold=True, color="FF0000")
            ws['D5'].number_format = '#,##0 " VNĐ"' # Định dạng tiền tệ
            
            # 3. Ghi Chi tiết Hóa đơn (Phần Table)
            start_row = 7
            ws.merge_cells(f'A{start_row}:F{start_row}') 
            ws[f'A{start_row}'] = "CHI TIẾT CÁC SẢN PHẨM"
            ws[f'A{start_row}'].font = openpyxl.styles.Font(bold=True, size=12)
            ws[f'A{start_row}'].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Tiêu đề bảng
            header = ['STT', 'Mã Tivi', 'Tên Tivi', 'Số Lượng', 'Đơn Giá (VNĐ)', 'Thành Tiền (VNĐ)']
            ws.append(header)
            
            # Định dạng tiêu đề
            header_row = ws[start_row + 1]
            fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for cell in header_row:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
            # Ghi dữ liệu
            current_row = start_row + 2
            # current_cart_data là: (ma_tivi, ten_tivi, so_luong, don_gia, thanh_tien)
            for i, item in enumerate(current_cart_data, 1):
                ma_tivi, ten_tivi, so_luong, don_gia, thanh_tien = item
                
                ws.cell(row=current_row, column=1, value=i) # STT
                ws.cell(row=current_row, column=2, value=ma_tivi).alignment = openpyxl.styles.Alignment(horizontal='center')
                ws.cell(row=current_row, column=3, value=ten_tivi) # Tên Tivi
                ws.cell(row=current_row, column=4, value=so_luong).alignment = openpyxl.styles.Alignment(horizontal='center')
                ws.cell(row=current_row, column=5, value=don_gia).number_format = '#,##0' # Đơn giá
                ws.cell(row=current_row, column=6, value=thanh_tien).number_format = '#,##0' # Thành tiền
                current_row += 1
                
            # Tự động điều chỉnh độ rộng cột
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        length = len(str(cell.value))
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), length))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value + 2 

            # 4. Lưu file
            wb.save(file_path)
            return True

        except Exception as e:
            messagebox.showerror("Lỗi Xuất Excel", f"Lỗi khi xuất file Excel: {e}")
            return False
    
    def export_last_invoice_from_variable(export_last_btn):
        global last_invoice_info
        
        if last_invoice_info:
            ma_hd, ma_kh, ma_nv, tong_tien, current_cart_data = last_invoice_info
            
            if export_invoice_to_excel(ma_hd, ma_kh, ma_nv, tong_tien, current_cart_data):
                # Vô hiệu hóa nút sau khi xuất thành công
                export_last_btn.config(state=tk.DISABLED)
        else:
            messagebox.showwarning("Không tìm thấy", "Chưa có hóa đơn nào được lập thành công trong phiên này.")
            
    # --- CÁC HÀM XỬ LÝ (Điều chỉnh) ---
    
    # ... (các hàm load_combobox_data, on_tivi_select, update_cart_tree, add_to_cart, remove_from_cart không thay đổi nhiều) ...
    # (Để giữ code ngắn gọn, tôi chỉ hiển thị những phần thay đổi quan trọng nhất)
    # Tuy nhiên, cần đảm bảo các tham số đã được bổ sung ở các hàm add_to_cart và remove_from_cart trước đó vẫn còn.

    def load_combobox_data():
        # Load dữ liệu cho các Combobox
        global tivi_list, customer_list, staff_list
        try:
            conn = connect_db()
            cursor = conn.cursor()

            # Load Khach hàng
            cursor.execute("SELECT ma_kh, ho_ten FROM khachhang ORDER BY ho_ten")
            customer_list = cursor.fetchall()
            customer_combo['values'] = [f"{kh[1]} (ID: {kh[0]})" for kh in customer_list]

            # Load Nhân viên
            cursor.execute("SELECT ma_nv, ho_ten FROM nhanvien ORDER BY ho_ten")
            staff_list = cursor.fetchall()
            staff_combo['values'] = [f"{nv[1]} (ID: {nv[0]})" for nv in staff_list]

            # Load Tivi (chỉ Tivi còn hàng)
            cursor.execute("SELECT ma_tivi, ten_tivi, gia, so_luong FROM tivi WHERE so_luong > 0 ORDER BY ten_tivi")
            tivi_list = cursor.fetchall()
            tivi_combo['values'] = [f"{tv[1]} (Giá: {tv[2]:,.0f})" for tv in tivi_list]
            
            if customer_combo['values']: customer_combo.current(0)
            if staff_combo['values']: staff_combo.current(0)
            
            # Reset Tivi combo và label nếu danh sách Tivi trống
            if tivi_combo['values']: 
                tivi_combo.current(0)
            else:
                tivi_combo.set("Hết hàng")
                tivi_combo['values'] = []
                ton_kho_label.config(text="0 cái")
                don_gia_label.config(text="0 VNĐ")
    
        except Exception as e:
            messagebox.showerror("Lỗi CSDL", f"Không thể tải dữ liệu: {e}")
        finally:
            if conn: conn.close()
            
    def on_tivi_select(event):
        # Cập nhật nhãn tồn kho và đơn giá khi chọn Tivi
        selected_index = tivi_combo.current()
        if selected_index < 0 or not tivi_list: 
             ton_kho_label.config(text="0 cái")
             don_gia_label.config(text="0 VNĐ")
             return
        
        selected_tivi = tivi_list[selected_index]
        so_luong_ton = selected_tivi[3]
        don_gia = selected_tivi[2]
        
        ton_kho_label.config(text=f"{so_luong_ton} cái")
        don_gia_label.config(text=f"{don_gia:,.0f} VNĐ")

    def update_cart_tree(cart_tree, total_label):
        # Cập nhật TreeView giỏ hàng và tổng tiền
        for item in cart_tree.get_children():
            cart_tree.delete(item)
        
        tong_tien = 0
        for item in current_cart:
            formatted_item = (item[0], item[1], item[2], f"{item[3]:,.0f}", f"{item[4]:,.0f}")
            cart_tree.insert('', 'end', values=formatted_item)
            tong_tien += item[4]
            
        total_label.config(text=f"{tong_tien:,.0f} VNĐ")

    def add_to_cart(tivi_combo, so_luong_entry, ton_kho_label, don_gia_label, cart_tree, total_label):
        selected_index = tivi_combo.current()
        if selected_index < 0 or not tivi_list:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn một Tivi")
            return
            
        try:
            so_luong_mua = int(so_luong_entry.get())
            if so_luong_mua <= 0:
                raise ValueError("Số lượng phải lớn hơn 0")
        except ValueError:
            messagebox.showwarning("Sai số lượng", "Vui lòng nhập số lượng mua hợp lệ")
            return

        selected_tivi = tivi_list[selected_index]
        ma_tivi = selected_tivi[0]
        ten_tivi = selected_tivi[1]
        don_gia = selected_tivi[2]
        so_luong_ton = selected_tivi[3]

        if so_luong_mua > so_luong_ton:
            messagebox.showwarning("Hết hàng", f"Số lượng tồn kho không đủ (Chỉ còn {so_luong_ton} cái)")
            return
            
        found = False
        for i, item in enumerate(current_cart):
            if item[0] == ma_tivi:
                new_sl = item[2] + so_luong_mua
                if new_sl > so_luong_ton:
                    messagebox.showwarning("Hết hàng", f"Tổng số lượng mua ({new_sl}) vượt quá tồn kho ({so_luong_ton})")
                    return
                thanh_tien = new_sl * don_gia
                current_cart[i] = (ma_tivi, ten_tivi, new_sl, don_gia, thanh_tien)
                found = True
                break
        
        if not found:
            thanh_tien = so_luong_mua * don_gia
            current_cart.append((ma_tivi, ten_tivi, so_luong_mua, don_gia, thanh_tien))

        update_cart_tree(cart_tree, total_label)
        so_luong_entry.delete(0, 'end')

    def remove_from_cart(cart_tree, total_label):
        selected_item = cart_tree.focus()
        if not selected_item:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn Tivi cần xóa khỏi giỏ")
            return
        
        values = cart_tree.item(selected_item, 'values')
        ma_tivi = values[0]
        
        for i, item in enumerate(current_cart):
            if str(item[0]) == str(ma_tivi):
                del current_cart[i]
                break
        
        update_cart_tree(cart_tree, total_label)

    def checkout(customer_combo, staff_combo, cart_tree, total_label, export_last_btn):
        global last_invoice_info # Sử dụng biến global
        
        if not current_cart:
            messagebox.showwarning("Giỏ hàng rỗng", "Vui lòng thêm sản phẩm vào giỏ")
            return
        
        try:
            ma_kh = customer_list[customer_combo.current()][0]
            ma_nv = staff_list[staff_combo.current()][0]
        except IndexError:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng chọn khách hàng và nhân viên")
            return

        if not messagebox.askyesno("Xác nhận", "Bạn có chắc muốn thanh toán và lưu hóa đơn này?"):
            return
        
        # Lưu lại giỏ hàng hiện tại (chưa reset) để lưu vào biến global và xuất Excel
        invoice_items_for_export = list(current_cart) 

        conn = None
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # --- Bắt đầu Transaction ---
            conn.start_transaction()

            tong_tien = sum(item[4] for item in invoice_items_for_export)
            ngay_lap = date.today()
            
            #  Tạo Hóa đơn (hoadon)
            sql_hoadon = "INSERT INTO hoadon (ma_nv, ma_kh, ngay_lap, tong_tien) VALUES (%s, %s, %s, %s)"
            cursor.execute(sql_hoadon, (ma_nv, ma_kh, ngay_lap, tong_tien))
            ma_hd_moi = cursor.lastrowid # Lấy ID của hóa đơn vừa tạo

            #  Thêm các Chi tiết Hóa đơn (chitiet_hoadon) và Cập nhật tồn kho
            sql_chitiet = "INSERT INTO chitiet_hoadon (ma_hd, ma_tivi, so_luong_mua, don_gia) VALUES (%s, %s, %s, %s)"
            sql_update_tivi = "UPDATE tivi SET so_luong = so_luong - %s WHERE ma_tivi = %s"
            
            for item in invoice_items_for_export:
                ma_tivi = item[0]
                so_luong_mua = item[2]
                don_gia = item[3]
                
                cursor.execute(sql_chitiet, (ma_hd_moi, ma_tivi, so_luong_mua, don_gia))
                cursor.execute(sql_update_tivi, (so_luong_mua, ma_tivi))

            # --- Kết thúc Transaction (lưu) ---
            conn.commit()
            
            messagebox.showinfo("Thành công", f"Đã lưu hóa đơn (Mã HĐ: {ma_hd_moi}) thành công!")
            
            # --- LƯU THÔNG TIN HÓA ĐƠN VỪA LẬP VÀ KÍCH HOẠT NÚT EXCEL ---
            last_invoice_info = (ma_hd_moi, ma_kh, ma_nv, tong_tien, invoice_items_for_export)
            export_last_btn.config(state=tk.NORMAL)
            
            # Reset giỏ hàng và giao diện
            current_cart.clear()
            update_cart_tree(cart_tree, total_label)
            load_combobox_data() 

        except Exception as e:
            
            if conn:
                conn.rollback()
            messagebox.showerror("Lỗi Giao dịch", f"Không thể lưu hóa đơn: {e}")
            last_invoice_info = None # Đảm bảo không lưu thông tin lỗi
            export_last_btn.config(state=tk.DISABLED)
        finally:
            if conn:
                conn.close()

    # --- Gán sự kiện và Tải dữ liệu ---
    tivi_combo.bind('<<ComboboxSelected>>', on_tivi_select)
    load_combobox_data()
    # Kích hoạt sự kiện lần đầu
    on_tivi_select(None)
    
    return frame